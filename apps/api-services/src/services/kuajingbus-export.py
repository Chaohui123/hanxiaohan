#!/usr/bin/env python3
"""跨境巴士订单导出 — 基于原始模板追加数据，100%保留格式"""
import sys, json, os
from datetime import datetime
from openpyxl import load_workbook

# Template is at /app/assets/ in Docker, at assets/ relative to project root on dev
script_dir = os.path.dirname(os.path.abspath(__file__))
template_path = os.path.join(script_dir, "..", "..", "..", "assets", "kuajingbus-template.xlsx")
if not os.path.exists(template_path):
    # Docker path
    template_path = "/app/assets/kuajingbus-template.xlsx"
input_json = sys.stdin.read()
data = json.loads(input_json)

# Load template
wb = load_workbook(template_path)
ws = wb[wb.sheetnames[0]]

# Find "请在此行开始填写真实订单" marker
start_row = 7  # fallback
for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=1):
    cell = row[0]
    if cell.value and "真实订单" in str(cell.value):
        start_row = cell.row + 1
        break

# Append data rows
ri = start_row
for order in data:
    vals = [
        "",                  # A 注释
        "1052",              # B 仓库代码
        "10",                # C 服务代码
        order.get("posting", ""),   # D 平台订单号
        order.get("tracking", ""),  # E 面单条形码
        "",                  # F 产品图片
        "",                  # G 产品名称
        str(order.get("qty", 1)),   # H 产品数量
        "1688",              # I 打包来源
        "",                  # J 快递单号
        str(order.get("weight", "")),  # K 预估重量
        "1688",              # L 采购平台
        order.get("id", ""), # M 采购单号
    ]
    for ci, val in enumerate(vals):
        ws.cell(row=ri, column=ci + 1, value=val)
    ri += 1

# Write to stdout as binary
import io
buf = io.BytesIO()
wb.save(buf)
sys.stdout.buffer.write(buf.getvalue())
